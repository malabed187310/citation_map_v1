import openpyxl
import requests
import pandas as pd
from PyQt5.QtWidgets import QApplication, QInputDialog

app = QApplication([])

query, ok_pressed = QInputDialog.getText(None, 'User Input', 'Enter article Name: ')

# Process the user input
if ok_pressed:
    print("User input:", query)
else:
    print("No input provided.")


# Set up the SERP API request URL and parameters
api_key = '3bec768da518b4b2c6090dbd1ea2331dbc66acb397ddaad6fc447be6f5cc7b55'
#query = 'Design, Integration, Testing and Lessons Learned of a Utility Operated Microgrid Master Controller'
url = f'https://serpapi.com/search?engine=google_scholar&q={query}&api_key={api_key}'

# Send the request and retrieve the search results
response = requests.get(url)
data = response.json()

# Extract the cited articles from the search results
list1 = []
if 'organic_results' in data:
    for result in data['organic_results']:
        # Extract cited articles if available
        if 'inline_links' in result:
            result2 = result['inline_links']
            if 'cited_by' in result2:
                result3 = result2['cited_by']
                list1.append(result3['serpapi_scholar_link'])


# Open a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Rename column
column_names = ['Cited Articles']
worksheet.append(column_names)

# Loop through each cited article and add it to the worksheet
for article in list1:
    worksheet.append([article])

# Save the Excel workbook
workbook.save('api_link.xlsx')


##########################################################################################################


# Read file with URL cited links
df = pd.read_excel('api_link.xlsx')

# Accessing the data points in the 'Cited Articles' column
cited_links = df['Cited Articles']

# Open a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Total number of pages to retrieve
num_pages = 8

# Loop through each cited link
list2 = []
for link in cited_links:
    # Define the URL for the current cited link
    endpoint = link + f'&api_key={api_key}'

    # Loop through each page of search results
    #current_page = 0
    #more_pages = True
    #while more_pages:
    for page in range(num_pages):
        url = f'{endpoint}&start={page*10}'

        # Send the request and retrieve the search results
        response = requests.get(url)
        data = response.json()

        # Extract the titles and authors of the search results
        if 'organic_results' in data:
            for result in data['organic_results']:
                title = result['title']

                ''' 
                # This finds all authors
                # Extract author and affiliation information if available
                if 'publication_info' in result:
                    author_info = result['publication_info']
                    if 'summary' in author_info:
                        authors = author_info['summary']
                    else:
                        authors = ''
                else:
                    authors = ''
                '''
                # This finds the first author
                # Extract author and affiliation information if available
                if 'publication_info' in result:
                    author_info = result['publication_info']
                    if 'authors' in author_info:
                        authors = author_info['authors'][0]
                        if 'name' in authors:
                            first_author = authors['name']
                            author_id = authors['author_id']

                            list2.append({'Title': title, 'Author': first_author, 'Author ID': author_id})
'''
if 'pagination' in data:
    next_page = data['pagination']
    if 'next' != None:
        current_page += 1
    else:
        more_pages = False
'''


column_names = ['Title', 'Author', 'Author ID']
worksheet.append(column_names)

for i in list2:
    worksheet.append([i['Title'], i['Author'], i['Author ID']])

# Save the Excel workbook
workbook.save('api_author.xlsx')


##############################################################################################


# Read file with URL cited links
df = pd.read_excel('api_author.xlsx')

# Accessing the data points in the 'Cited Articles' column
author_id = df['Author ID']

# Open a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

list3 = []
for result in author_id:
    author_url = f'https://serpapi.com/search.json?engine=google_scholar_author&author_id={result}&api_key={api_key}'

    # Send the request and retrieve the search results
    response = requests.get(author_url)
    data = response.json()

    if 'author' in data:
        info = data['author']
        if 'name' in info:
            names = info['name']
            
            
        if 'affiliations' in info:
            location = info['affiliations']
            
        
        list3.append({'Author': names, 'University': location})


# Renames columns
column_names = ['Author', 'University']
worksheet.append(column_names)

# Loop through each title in the title_info list and add it to the worksheet
for i in list3:
    worksheet.append([i['Author'], i['University']])
             

# Save the Excel workbook
workbook.save('api_location.xlsx')


#############################################################################################


app.quit()